"""
Step-0 Agent Classification Engine (7×9 Matrix Cell Recommendation) — v1.3.3
=========================================================================

Enhancements in v1.2
-------------------
1) Negative regex rules (penalties)
   - Rulebook has Rule_Mode = POS or NEG.
   - NEG rules subtract from the label score (auditable evidence recorded).

2) Phrase proximity logic
   - Rulebook supports:
        Proximity_Terms (two regex groups separated by '||' or legacy '(A)|(B)')
        Proximity_Window_Chars (integer)
     If terms from group A occur within +/- window chars of terms from group B,
     the rule fires (POS or NEG).

3) Confidence threshold warnings in Excel
   - Implemented in the XLSX template using formulas + conditional formatting:
       Confidence_Cell < 0.55  -> HIGH
       0.55..0.75              -> MEDIUM
       >= 0.75                 -> OK
   - Script writes recommendations; warnings auto-calc in Excel.

Human approval is ALWAYS required
---------------------------------
The script never writes *_Final fields.
Finalize mode only produces Step-2-ready output when ALL agents are APPROVED
and *_Final fields are filled by a human.

"""

import argparse
import json
import math
from pdb import run
import re
from collections import Counter, defaultdict
from datetime import datetime
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook


# -----------------------------
# Utility: text normalization
# -----------------------------
def norm_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def tokenize(s: str) -> List[str]:
    s = norm_text(s)
    return re.findall(r"[a-z0-9]+", s)

# --- New Agent_Metadata (functional-profile) layout support ---
FP_DIMS = [
    "Agent Name", "Business Objective", "Primary Role", "Detailed Micro Functionalities",
    "Trigger / Invocation", "Inputs Consumed", "Processing Activities", "Primary Outputs",
    "Decision-Making Responsibility", "Human-in-the-Loop (HITL)", "Knowledge Sources Used",
    "Tool & System Interactions", "Collaboration Pattern", "SDLC / Business Lifecycle Phase",
    "Complexity Characteristics", "Autonomy Level", "Governance & Compliance Needs",
    "Error Handling & Recovery", "Execution Frequency / Runtime Pattern", "Expected Business Value",
]
CAPABILITY_DIMS = [
    "Primary Role", "Detailed Micro Functionalities", "Processing Activities",
    "Primary Outputs", "Inputs Consumed", "Tool & System Interactions",
    "SDLC / Business Lifecycle Phase", "Trigger / Invocation",
]


def find_header_row(ws, key="Agent_ID", max_scan=8):
    for i in range(1, max_scan + 1):
        vals = [str(c.value).strip() if c.value is not None else "" for c in ws[i]]
        if key in vals:
            return i
    return 1


def detect_meta_layout(ws):
    r1c1 = str(ws.cell(1, 1).value or "").strip()
    r2c2 = str(ws.cell(2, 2).value or "").strip()
    if r1c1 == "Agent_ID" and r2c2 == "Agent Name":
        return "v3_profile"
    return "legacy"


def _flatten_dims(row):
    return "\n".join(f"{d}: {' '.join(str(row.get(d, '')).split())}"
                     for d in FP_DIMS if str(row.get(d, '')).strip())


def capability_signal_text(agent_description):
    """Capability-only text (flat descriptions pass through)."""
    if not agent_description:
        return ""
    labels = sorted(FP_DIMS, key=len, reverse=True)
    alt = "|".join(re.escape(l) for l in labels)
    pat = re.compile(rf"(?:^|\n)\s*({alt})\s*:\s*(.*?)(?=\n\s*(?:{alt})\s*:|\Z)", re.S)
    dims = {m.group(1).strip(): " ".join(m.group(2).split()).strip() for m in pat.finditer(agent_description)}
    if not dims:
        return agent_description
    return " ".join(dims.get(d, "") for d in CAPABILITY_DIMS if dims.get(d)).strip()


def build_meta_v3(ws):
    """Read the functional-profile Agent_Metadata: Agent_ID row1/col1, 20 dims row2,
    recommendation columns headed in row1, data from row3."""
    dims = [ws.cell(2, c).value for c in range(2, 22)]
    meta_by_id = {}
    for r in range(3, ws.max_row + 1):
        aid = ws.cell(r, 1).value
        if aid is None or str(aid).strip() == "":
            continue
        aid = str(aid).strip()
        row = {dims[j]: (ws.cell(r, 2 + j).value or '') for j in range(len(dims))}
        meta_by_id[aid] = {"Agent_Name": str(row.get("Agent Name", aid)).strip(),
                           "Agent_Description": _flatten_dims(row), "_row": r}
    col_index = {"Agent_ID": 1}
    for c in range(2, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h and str(h).strip() and not str(h).startswith("Agent Functional Profile"):
            col_index[str(h).strip()] = c
    return meta_by_id, col_index, 3



def cosine_sim(a: Counter, b: Counter) -> float:
    if not a or not b:
        return 0.0
    dot = 0.0
    for k, va in a.items():
        dot += va * b.get(k, 0)
    na = math.sqrt(sum(v*v for v in a.values()))
    nb = math.sqrt(sum(v*v for v in b.values()))
    if na == 0 or nb == 0:
        return 0.0
    return float(dot / (na * nb))


# --------------------------------------
# Semantic scoring HOOK (replaceable)
# --------------------------------------
def semantic_score(agent_text: str, catalog_text: str) -> float:
    """
    Default semantic scoring uses token cosine similarity (bag-of-words).
    Replace with embeddings later if desired.
    Must return a float in [0,1].
    """
    a = Counter(tokenize(agent_text))
    b = Counter(tokenize(catalog_text))
    return cosine_sim(a, b)


# -----------------------------
# Read sheet helpers
# -----------------------------
def read_table(ws, header_row=1) -> Tuple[List[str], List[Dict]]:
    headers = [str(c.value).strip()
               if c.value is not None else "" for c in ws[header_row]]
    rows = []
    for r in ws.iter_rows(min_row=header_row+1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None
        rows.append(d)
    return headers, rows


def index_by(rows: List[Dict], key: str) -> Dict:
    out = {}
    for r in rows:
        k = r.get(key)
        if k is None or str(k).strip() == "":
            continue
        out[str(k).strip()] = r
    return out


# -----------------------------
# Proximity matcher
# -----------------------------
def proximity_match(text: str, group_a_regex: str, group_b_regex: str, window_chars: int) -> Optional[Dict]:
    """
    Returns evidence dict if any match of group A is within +/- window_chars of group B.
    Uses character offsets in the normalized compiled agent_text (not per-field).
    """
    if window_chars <= 0:
        return None

    # Find all occurrences of each group
    a_iter = list(re.finditer(group_a_regex, text, flags=re.IGNORECASE))
    b_iter = list(re.finditer(group_b_regex, text, flags=re.IGNORECASE))
    if not a_iter or not b_iter:
        return None

    # Check distance between any a and b
    for ma in a_iter:
        for mb in b_iter:
            dist = abs(ma.start() - mb.start())
            if dist <= window_chars:
                # Provide compact evidence excerpt around midpoint
                lo = max(0, min(ma.start(), mb.start()) - 40)
                hi = min(len(text), max(ma.end(), mb.end()) + 40)
                excerpt = text[lo:hi]
                return {
                    "a_match": ma.group(0),
                    "b_match": mb.group(0),
                    "distance_chars": dist,
                    "excerpt": excerpt
                }
    return None


# -----------------------------
# Rule engine (POS/NEG + Proximity)
# -----------------------------
def apply_rules(agent_text: str, rule_rows: List[Dict]) -> Tuple[Dict[str, float], Dict[str, float], List[Dict]]:
    """
    Extended in v1.4 to support SUPER_GENERIC rule types:
    - INTENT  -> Engineering_Dimension boost
    - PLANE   -> Architecture_Domain boost
    - POSTURE -> Architecture_Domain boost
    """

    """
    Returns:
      arch_scores[label] = score
      eng_scores[label]  = score
      fired_rules = evidence list
    """
    arch_scores = defaultdict(float)
    eng_scores = defaultdict(float)
    fired = []

    for rr in rule_rows:
        rule_id = str(rr.get("Rule_ID", "")).strip()

        # Backward compatibility
        mode = str(rr.get("Rule_Mode", "POS") or "POS").strip().upper()
        if mode not in ("POS", "NEG"):
            mode = "POS"

        rule_type = str(rr.get("Rule_Type", "") or "").strip().upper()

        regex = rr.get("Regex", "") or ""
        # SUPER_GENERIC uses Rule_Logic as the regex pattern
        if (not regex) and rr.get("Rule_Logic"):
            regex = str(rr.get("Rule_Logic") or "")
        prox_terms = rr.get("Proximity_Terms", "") or ""
        prox_window = rr.get("Proximity_Window_Chars", "") or ""

        # SUPER_GENERIC schema
        target = str(rr.get("Target_Dimension", rr.get("Target", ""))).strip()
        label = str(rr.get("Target_Label", rr.get("Label", ""))).strip()

        weight = float(rr.get("Weight", 0) or 0)
        rationale = str(rr.get("Rationale", "")).strip()

        if not label or weight <= 0 or target not in ("Architecture_Domain", "Engineering_Dimension"):
            continue
        if mode not in ("POS", "NEG"):
            mode = "POS"

        fired_now = False
        evidence = {"type": "", "match_excerpt": ""}

        # 1) Proximity rule if configured (takes precedence)
        if prox_terms and prox_window:
            # Expect two groups separated by '|' at top-level; we store as "(A)|(B)" in template.
            # We'll split on ')|(' robustly by extracting two groups:
            # If user keeps "A|B" without parentheses, we treat the first '|' as separator.
            pt = str(prox_terms).strip()

            # Proximity_Terms supported formats:
            # 1) Canonical: <A> || <B>
            # 2) Legacy: (A)|(B) (including extra wrapping like ((A))|((B)))
            # Anything else is rejected to avoid ambiguous splitting.
            def split_proximity_terms(pt_raw: str) -> Tuple[str, str]:
                pt_raw = (pt_raw or "").strip()
                if not pt_raw:
                    return "", ""

                # Canonical delimiter (recommended)
                if "||" in pt_raw:
                    a, b = [p.strip() for p in pt_raw.split("||", 1)]
                    return a, b

                # Legacy: (A)|(B) with optional whitespace; DOTALL so inner groups can contain pipes
                mm = re.match(r"^\s*\((.*)\)\s*\|\s*\((.*)\)\s*$",
                              pt_raw, flags=re.DOTALL)
                if mm:
                    a, b = mm.group(1).strip(), mm.group(2).strip()

                    # If people wrote extra wrapping like ((A))|((B)), strip ONE layer of outer parens safely
                    def strip_outer(x: str) -> str:
                        x = x.strip()
                        if x.startswith("(") and x.endswith(")"):
                            return x[1:-1].strip()
                        return x

                    return strip_outer(a), strip_outer(b)

                raise ValueError(
                    f"Invalid Proximity_Terms format for Rule_ID={rule_id}. "
                    f"Expected '<A> || <B>' or legacy '(A)|(B)'. Got: {pt_raw}"
                )

            a_rgx, b_rgx = split_proximity_terms(pt)

            # Regex pre-validation (clear errors)
            try:
                re.compile(a_rgx)
                re.compile(b_rgx)
            except re.error as e:
                raise ValueError(
                    f"Invalid proximity regex for Rule_ID={rule_id}: {str(e)}. A='{a_rgx}' B='{b_rgx}'"
                )
            try:
                window = int(float(prox_window))
            except (TypeError, ValueError):
                window = 0

            prox_ev = proximity_match(agent_text, a_rgx, b_rgx, window)
            if prox_ev:
                fired_now = True
                evidence["type"] = "PROXIMITY"
                evidence["match_excerpt"] = prox_ev["excerpt"][:200]
                evidence["proximity"] = prox_ev

        # 2) Regex rule
        if not fired_now and regex:
            try:
                m = re.search(regex, agent_text, flags=re.IGNORECASE)
            except re.error as e:
                raise ValueError(
                    f"Invalid Regex pattern for Rule_ID={rule_id}: {str(e)}. Regex='{regex}'")
            if m:
                fired_now = True
                evidence["type"] = "REGEX"
                evidence["match_excerpt"] = m.group(0)[:160]

        if not fired_now:
            continue

        signed_weight = weight if mode == "POS" else -weight

        fired.append({
            "rule_id": rule_id,
            "rule_mode": mode,
            "target": target,
            "label": label,
            "weight": weight,
            "signed_weight": signed_weight,
            "evidence_type": evidence["type"],
            "match_excerpt": evidence.get("match_excerpt", ""),
            "proximity": evidence.get("proximity", None),
            "rationale": rationale
        })

        if target == "Architecture_Domain":
            arch_scores[label] += signed_weight
        else:
            eng_scores[label] += signed_weight

    # Prevent negative totals from making confidence odd:
    # shift scores so the minimum is 0 if needed (auditable transform)
    def shift_nonnegative(scores: Dict[str, float]) -> Tuple[Dict[str, float], float]:
        if not scores:
            return scores, 0.0
        mn = min(scores.values())
        if mn >= 0:
            return scores, 0.0
        shift = -mn
        shifted = {k: v + shift for k, v in scores.items()}
        return shifted, shift

    arch_scores, arch_shift = shift_nonnegative(dict(arch_scores))
    eng_scores, eng_shift = shift_nonnegative(dict(eng_scores))
    if arch_shift or eng_shift:
        fired.append({
            "rule_id": "SCORE_SHIFT",
            "rule_mode": "INFO",
            "target": "INTERNAL",
            "label": "",
            "weight": 0,
            "signed_weight": 0,
            "evidence_type": "SHIFT",
            "match_excerpt": "",
            "proximity": None,
            "rationale": f"Shifted scores to be non-negative for confidence calc. arch_shift={arch_shift}, eng_shift={eng_shift}"
        })

    return arch_scores, eng_scores, fired


def pick_top(scores: Dict[str, float]) -> Tuple[str, float, List[Tuple[str, float]]]:
    if not scores:
        return ("", 0.0, [])
    ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    top_label, top_score = ranked[0]
    total = sum(v for _, v in ranked) or 1.0
    conf = float(top_score / total)
    return (top_label, conf, ranked[:5])


# -----------------------------
# Cell recommendation
# -----------------------------
def recommend_cell(agent_text: str, cells: List[Dict]) -> Tuple[Dict, List[Dict]]:
    scored = []
    for c in cells:
        catalog_text = f"{c.get('Architecture_Domain', '')} {c.get('Engineering_Dimension', '')} {c.get('Cell_Description', '')}"
        s = semantic_score(agent_text, catalog_text)
        scored.append({**c, "semantic_score": float(s)})
    scored.sort(key=lambda x: x["semantic_score"], reverse=True)
    best = scored[0] if scored else {}
    top3 = scored[:3]
    if len(scored) >= 2:
        gap = max(
            0.0, min(1.0, best["semantic_score"] - scored[1]["semantic_score"] + 0.5))
    else:
        gap = best.get("semantic_score", 0.0)
    best["semantic_confidence"] = float(max(0.0, min(1.0, gap)))
    return best, top3


def run_logic(args):
    wb = load_workbook(args.in_xlsx)

    for s in ["Agent_Metadata", "Classification_Inputs", "Cell_Catalog"]:
        if s not in wb.sheetnames:
            raise RuntimeError(f"Missing required sheet: {s}")

    wb = load_workbook(args.in_xlsx)

    for s in ["Agent_Metadata", "Classification_Inputs", "Cell_Catalog"]:
        if s not in wb.sheetnames:
            raise RuntimeError(f"Missing required sheet: {s}")

    # Rulebook sheet can be toggled
    if args.rulebook_sheet not in wb.sheetnames:
        raise RuntimeError(f"Missing rulebook sheet: {args.rulebook_sheet}")
    # (legacy compatibility note)

    _ws_meta0 = wb["Agent_Metadata"]
    _meta_layout = detect_meta_layout(_ws_meta0)
    if _meta_layout == "v3_profile":
        meta_by_id, _v3_col_index, _v3_data_start = build_meta_v3(_ws_meta0)
    else:
        _hr = find_header_row(_ws_meta0)
        _, meta_rows = read_table(_ws_meta0, header_row=_hr)
        meta_by_id = index_by(meta_rows, "Agent_ID")
        _v3_col_index, _v3_data_start = None, None
    _, input_rows = read_table(wb["Classification_Inputs"], header_row=1)
    _, rule_rows = read_table(wb[args.rulebook_sheet], header_row=1)
    _, cell_rows = read_table(wb["Cell_Catalog"], header_row=1)
    inputs_by_id = index_by(input_rows, "Agent_ID")

    audit = {
        "run_ts": datetime.utcnow().isoformat() + "Z",
        "in_xlsx": args.in_xlsx,
        "version": "v1.4.1",
        "agents": []
    }

    ws_meta = wb["Agent_Metadata"]
    if _meta_layout == "v3_profile":
        col_index = _v3_col_index
        _data_start = _v3_data_start
    else:
        _hr2 = find_header_row(ws_meta)
        meta_headers = [str(c.value).strip()
                        if c.value is not None else "" for c in ws_meta[_hr2]]
        col_index = {h: i+1 for i, h in enumerate(meta_headers) if h}
        _data_start = _hr2 + 1

    def set_cell(row_idx: int, col_name: str, value):
        if col_name not in col_index:
            return
        ws_meta.cell(row=row_idx, column=col_index[col_name]).value = value

    for r_idx in range(_data_start, ws_meta.max_row + 1):
        agent_id = ws_meta.cell(
            row=r_idx, column=col_index.get("Agent_ID", 1)).value
        if agent_id is None or str(agent_id).strip() == "":
            continue
        agent_id = str(agent_id).strip()

        meta = meta_by_id.get(agent_id, {})
        inp = inputs_by_id.get(agent_id, {})

        agent_text_parts = []
        for k in ["Agent_Name", "Agent_Description"]:
            v = meta.get(k)
            if v is not None:
                if k == "Agent_Description":
                    v = capability_signal_text(str(v))
                agent_text_parts.append(str(v))
        for k, v in inp.items():
            if k == "Agent_ID":
                continue
            if v is None or str(v).strip() == "":
                continue
            agent_text_parts.append(f"{k}:{str(v).replace('|', ' ')}")
        agent_text = " | ".join(agent_text_parts)

        arch_scores, eng_scores, fired = apply_rules(agent_text, rule_rows)
        arch_top, arch_conf, arch_ranked = pick_top(arch_scores)
        eng_top, eng_conf, eng_ranked = pick_top(eng_scores)

        best_cell, top3 = recommend_cell(agent_text, cell_rows)

        arch_rec = arch_top or best_cell.get("Architecture_Domain", "")
        eng_rec = eng_top or best_cell.get("Engineering_Dimension", "")

        # Choose a semantic cell consistent with rule-rec, else best semantic
        cell_rec = best_cell.get("Cell_ID", "")
        cell_conf = float(best_cell.get("semantic_score", 0.0))
        for c in top3:
            if c.get("Architecture_Domain") == arch_rec and c.get("Engineering_Dimension") == eng_rec:
                cell_rec = c.get("Cell_ID", cell_rec)
                cell_conf = float(c.get("semantic_score", cell_conf))
                break

        sem_conf = float(best_cell.get("semantic_confidence", 0.0))
        conf_arch = float(arch_conf if arch_top else sem_conf)
        conf_eng = float(eng_conf if eng_top else sem_conf)
        conf_cell = float(cell_conf)

        evidence_rules = "; ".join(
            [f"{x['rule_id']}[{x['rule_mode']}]→{x['label']}({x['signed_weight']:+.2f})"
             + (f"[{x.get('match_excerpt', '')}]" if x.get("match_excerpt") else "")
             for x in fired if x.get("rule_id") != "SCORE_SHIFT"]
        )[:2000]
        evidence_sem = " | ".join(
            [f"{x.get('Cell_ID')}:{x.get('semantic_score'):.3f}" for x in top3])

        set_cell(r_idx, "Architecture_Domain_Recommended", arch_rec)
        set_cell(r_idx, "Engineering_Dimension_Recommended", eng_rec)
        set_cell(r_idx, "Cell_ID_Recommended", cell_rec)
        set_cell(r_idx, "Confidence_Architecture", round(conf_arch, 4))
        set_cell(r_idx, "Confidence_Engineering", round(conf_eng, 4))
        set_cell(r_idx, "Confidence_Cell", round(conf_cell, 4))
        set_cell(r_idx, "Evidence_Rules", evidence_rules)
        set_cell(r_idx, "Evidence_Semantic_Top3", evidence_sem)

        cur_status = ws_meta.cell(
            row=r_idx, column=col_index.get("Approval_Status", 12)).value
        if cur_status is None or str(cur_status).strip() == "":
            set_cell(r_idx, "Approval_Status", "PENDING")

        audit["agents"].append({
            "Agent_ID": agent_id,
            "agent_text_compiled": agent_text,
            "rule_engine": {
                "arch_scores": arch_scores,
                "eng_scores": eng_scores,
                "arch_ranked_top5": arch_ranked,
                "eng_ranked_top5": eng_ranked,
                "fired_rules": fired
            },
            "semantic": {
                "best_cell": {k: best_cell.get(k) for k in ["Cell_ID", "Architecture_Domain", "Engineering_Dimension", "semantic_score", "semantic_confidence"]},
                "top3_cells": [{k: c.get(k) for k in ["Cell_ID", "Architecture_Domain", "Engineering_Dimension", "semantic_score"]} for c in top3]
            },
            "recommendation": {
                "Architecture_Domain_Recommended": arch_rec,
                "Engineering_Dimension_Recommended": eng_rec,
                "Cell_ID_Recommended": cell_rec,
                "Confidence_Architecture": conf_arch,
                "Confidence_Engineering": conf_eng,
                "Confidence_Cell": conf_cell,
                "Approval_Status": (ws_meta.cell(row=r_idx, column=col_index.get("Approval_Status", 12)).value or "PENDING")
            }
        })

    out_xlsx = args.out_xlsx.strip() or args.in_xlsx
    wb.save(out_xlsx)

    with open(args.audit_json, "w", encoding="utf-8") as f:
        json.dump(audit, f, indent=2, ensure_ascii=False)

    if args.finalize:
        wb2 = load_workbook(out_xlsx)
        ws_m = wb2["Agent_Metadata"]
        if detect_meta_layout(ws_m) == "v3_profile":
            idx = {"Agent_ID": 1}
            for _c in range(2, ws_m.max_column + 1):
                _h = ws_m.cell(1, _c).value
                if _h and not str(_h).startswith("Agent Functional Profile"):
                    idx[str(_h).strip()] = _c
            _fin_start = 3
        else:
            _fhr = find_header_row(ws_m)
            headers = [str(c.value).strip()
                       if c.value is not None else "" for c in ws_m[_fhr]]
            idx = {h: i+1 for i, h in enumerate(headers) if h}
            _fin_start = _fhr + 1

        def val(row, col): return ws_m.cell(row=row, column=idx[col]).value

        pending = []
        for r in range(_fin_start, ws_m.max_row + 1):
            aid = val(r, "Agent_ID")
            if aid is None or str(aid).strip() == "":
                continue
            status = str(val(r, "Approval_Status") or "").strip().upper()
            a_final = str(val(r, "Architecture_Domain_Final") or "").strip()
            e_final = str(val(r, "Engineering_Dimension_Final") or "").strip()
            if status != "APPROVED" or not a_final or not e_final:
                pending.append(str(aid).strip())
        if pending:
            raise RuntimeError("Finalize requested but some agents are not APPROVED and/or missing *_Final fields. "
                               f"Pending: {pending[:20]}{'...' if len(pending) > 20 else ''}")

        from openpyxl import Workbook as WB
        out = WB()
        ws_out = out.active
        ws_out.title = "Agent_Metadata"
        ws_out.append(["Agent_ID", "Architecture_Domain",
                      "Engineering_Dimension", "Cell_ID"])
        for r in range(2, ws_m.max_row + 1):
            aid = val(r, "Agent_ID")
            if aid is None or str(aid).strip() == "":
                continue
            ws_out.append([
                str(aid).strip(),
                str(val(r, "Architecture_Domain_Final")).strip(),
                str(val(r, "Engineering_Dimension_Final")).strip(),
                str(val(r, "Cell_ID_Final") or "").strip()
            ])
        out.save(args.step2_out_xlsx)

    print(f"Updated workbook: {out_xlsx}")
    print(f"Audit JSON: {args.audit_json}")
    if args.finalize:
        print(f"Step-2-ready workbook: {args.step2_out_xlsx}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in_xlsx", required=True, help="Step-0 classification workbook")
    ap.add_argument("--rulebook_sheet", default="Rulebook", help="Which rulebook tab to use (e.g., Rulebook or Rulebook_SUPER)")
    ap.add_argument("--out_xlsx", default="", help="Optional output xlsx path (otherwise overwrite input)")
    ap.add_argument("--audit_json", default="classification_audit.json", help="Audit JSON output")
    ap.add_argument("--finalize", action="store_true",
                    help="Create Step-2-ready workbook only when all agents are APPROVED and *_Final filled.")
    ap.add_argument("--step2_out_xlsx", default="Step2_ready_with_final_mapping.xlsx",
                    help="Output workbook produced on --finalize")
    args = ap.parse_args()
    run_logic(args)

# def main():
#     from argparse import Namespace
#     args = Namespace(
#         in_xlsx="C:/git-repos/aditya/remote_agent_testing/hello_world_agent/v1.9_2 (2).xlsx",
#         rulebook_sheet="RB_GOLD_ANCHOR_v1.7.6",
#         out_xlsx="output_a1.xlsx",
#         audit_json="output_json_a1.json",
#         finalize=False,
#     )
#     # run_logic(in_xlsx: str="C:/git-repos/aditya/remote_agent_testing/hello_world_agent/v1.9_2 (2).xlsx")
#     run_logic(args)


if __name__ == "__main__":
    main()
